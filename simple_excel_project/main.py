import openpyxl
from datetime import date
import os

# Sample data
invoices = {
    'Customer A': {1: {'Amount': 1000, 'DueDate': date(2023, 3, 15)}, 
                   2: {'Amount': 500, 'DueDate': date(2023, 3, 1)},
                   3: {'Amount': 2000, 'DueDate': date(2023, 2, 1)}},
    'Customer B': {4: {'Amount': 750, 'DueDate': date(2023, 3, 10)},
                   5: {'Amount': 1000, 'DueDate': date(2023, 2, 1)}}}

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AR Aging Report"

# Header row
ws['A1'] = "Customer Name"
ws['B1'] = "Invoice Number"
ws['C1'] = "Invoice Amount"
ws['D1'] = "Due Date"
ws['E1'] = "Days Overdue"

# Iterate through invoices and populate worksheet
row = 2  # Start on row 2
for customer, inv in invoices.items():
    for inv_num, inv_details in inv.items():
        inv_amt = inv_details['Amount']
        due_date = inv_details['DueDate']
        days_overdue = (date.today() - due_date).days
        if days_overdue > 0:
            ws.cell(row=row, column=1, value=customer)
            ws.cell(row=row, column=2, value=inv_num)
            ws.cell(row=row, column=3, value=inv_amt)
            ws.cell(row=row, column=4, value=due_date.strftime("%m/%d/%y"))
            ws.cell(row=row, column=5, value=days_overdue)
            row += 1

# Save workbook
wb.save('accounts_receivable_aging_report.xlsx')
os.system('start excel.exe "accounts_receivable_aging_report.xlsx"')
