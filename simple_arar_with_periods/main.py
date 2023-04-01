import openpyxl
from datetime import datetime
import os

# Create workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "AR Aging Report"

# Header row
ws['A1'] = "Customer Name"
ws['B1'] = "Invoice Number"
ws['C1'] = "Invoice Amount"
ws['D1'] = "Due Date"
ws['E1'] = "Days Overdue"

invoices = {
    'JAY': {
        1: {'Amount': 1000, 'DueDate': datetime(2023, 10, 22)}, 
        2: {'Amount': 500, 'DueDate': datetime(2023, 5, 15)},
        3: {'Amount': 2000, 'DueDate': datetime(2023, 6, 10)}
    },
    'CJ': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 5, 17)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 4, 21)}
    },
    'Kim': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 12, 27)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 8, 6)}
    },
    'Ghemar': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 5, 10)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 8, 4)}
    },
    'Vincent': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 4, 9)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 7, 4)}
    },
    'Belha': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 9, 29)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 9, 27)}
    },
    'Isabella': {
        4: {'Amount': 750, 'DueDate': datetime(2023, 7, 10)},
        5: {'Amount': 1000, 'DueDate': datetime(2023, 7, 1)}
    }
}

ws['F1'] = "0-30 days"
ws['G1'] = "31-60 days"
ws['H1'] = "61-90 days"
ws['I1'] = "91-120 days"
ws['J1'] = "Over 120 days"

invoice_total = 0
period_totals = [0, 0, 0, 0, 0]


row = 2  # Start on row 2
for customer, inv in invoices.items():
    for invoice_number, inv_details in inv.items():
        invoice_amount = inv_details['Amount']
        due_date = inv_details['DueDate']
        days_overdue = (due_date - datetime.today()).days
        if days_overdue >= 0:
            ws.cell(row=row, column=1, value=customer)
            ws.cell(row=row, column=2, value=invoice_number)
            ws.cell(row=row, column=3, value=invoice_amount)
            ws.cell(row=row, column=4, value=due_date.strftime("%m/%d/%y"))
            ws.cell(row=row, column=5, value=days_overdue)

            invoice_total += invoice_amount

            if days_overdue / 30 <= 1:
                ws.cell(row=row, column=6, value=invoice_amount)
                period_totals[0] += invoice_amount
            else:
                ws.cell(row=row, column=6, value=0)

            if days_overdue / 30 <= 2 and days_overdue / 30 > 1:
                ws.cell(row=row, column=7, value=invoice_amount)
                period_totals[1] += invoice_amount
            else:
                ws.cell(row=row, column=7, value=0)

            if days_overdue / 30 <= 3 and days_overdue / 30 >  2:
                ws.cell(row=row, column=8, value=invoice_amount)
                period_totals[2] += invoice_amount
            else:
                ws.cell(row=row, column=8, value=0)

            if days_overdue / 30 <= 4 and days_overdue / 30 >  3:
                ws.cell(row=row, column=9, value=invoice_amount)
                period_totals[3] += invoice_amount
            else:
                ws.cell(row=row, column=9, value=0)

            if days_overdue / 30 > 4:
                ws.cell(row=row, column=10, value=invoice_amount)
                period_totals[4] += invoice_amount
            else:
                ws.cell(row=row, column=10, value=0)
                
            row += 1


ws.cell(row=row, column=2, value="Invoice Total")
ws.cell(row=row, column=3, value=invoice_total)

for i in range(len(period_totals)):
    ws.cell(row=row, column=6+i, value=period_totals[i])

        
wb.save('accounts_receivable_aging_report.xlsx')
os.system('start excel.exe "accounts_receivable_aging_report.xlsx"')
