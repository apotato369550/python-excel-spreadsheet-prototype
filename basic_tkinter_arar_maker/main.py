import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import tkinter as tk
from tkinter import *
from tkinter import messagebox as mb

class ARARMaker(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)

        self.max_width = 900
        self.max_height = 575

        self.title("ARAR Maker V0.1")
        self.geometry(str(self.max_width) + "x" + str(self.max_height))
        self.resizable(False, False)

        self.invoices = {}
        self.period_totals = [0, 0, 0, 0, 0, 0]
        self.invoice_total = 0

        self.make_user_interface()
        self.create_workbook()
        self.load_invoices()
        self.update_invoices()
        self.update_statistics()
        self.mainloop()

        return

    def update_statistics(self):
        self.item_listbox.config(state=NORMAL)
        self.item_listbox.delete(1, END)
        self.item_listbox.insert(END, f"Total Amount Due: {self.invoice_total} \n")
        self.item_listbox.insert(END, f"\n")
        self.item_listbox.insert(END, f"Total Amount overdue (<0 days): {self.period_totals[0]} \n")
        self.item_listbox.insert(END, f"Total Amount due in 0-30 days: {self.period_totals[1]} \n")
        self.item_listbox.insert(END, f"Total Amount due in 31-60 days: {self.period_totals[2]} \n")
        self.item_listbox.insert(END, f"Total Amount due in 61-90 days: {self.period_totals[3]} \n")
        self.item_listbox.insert(END, f"Total Amount due in 91-120 days: {self.period_totals[4]} \n")
        self.item_listbox.insert(END, f"Total Amount due in >120 days: {self.period_totals[5]} \n")
        self.item_listbox.config(state=DISABLED)
        return

    def update_invoices(self):
        # test me
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "ARAR"

        self.headers = ["Customer Name", "Invoice Number", "Invoice Amount", "Due Date", "Days Overdue", "<0 days", "0-30 days", "31-60 days", "61-90 days", "91-120 days", "Over 120 days"]

        # Loop through the headers and set the values in the cells
        for index, header in enumerate(self.headers):
            self.worksheet.cell(row=1, column=index+1, value=header)
        
        row = 2
        for invoice_number, invoice in self.invoices.items():
            invoice_amount = float(invoice["amount"])
            customer_name = invoice["customer_name"]
            due_date = invoice["due_date"]
            print(due_date)
            print(datetime.today())
            days_overdue = (due_date - datetime.today()).days
            print("DAYS OVERDUE: " + str(days_overdue))

            self.invoice_total += invoice_amount

            self.worksheet.cell(row=row, column=1, value=customer_name)
            self.worksheet.cell(row=row, column=2, value=invoice_number)
            self.worksheet.cell(row=row, column=3, value=invoice_amount)
            self.worksheet.cell(row=row, column=4, value=due_date)
            self.worksheet.cell(row=row, column=5, value=days_overdue)

            if days_overdue < 0:
                self.worksheet.cell(row=row, column=6, value=invoice_amount)
                self.period_totals[0] += invoice_amount
                print("<0 DAYS")
            else:
                self.worksheet.cell(row=row, column=6, value=0)

            if days_overdue >= 0 and days_overdue <= 30:
                self.worksheet.cell(row=row, column=7, value=invoice_amount)
                self.period_totals[1] += invoice_amount
                print("0-30 DAYS")
            else:
                self.worksheet.cell(row=row, column=7, value=0)

            if days_overdue >= 31 and days_overdue <= 60:
                self.worksheet.cell(row=row, column=8, value=invoice_amount)
                self.period_totals[2] += invoice_amount
                print("31-60 DAYS")
            else:
                self.worksheet.cell(row=row, column=8, value=0)

            if days_overdue >= 61 and days_overdue <= 90:
                self.worksheet.cell(row=row, column=9, value=invoice_amount)
                self.period_totals[3] += invoice_amount
                print("61-90 DAYS")
            else:
                self.worksheet.cell(row=row, column=9, value=0)

            if days_overdue >= 91 and days_overdue <= 120:
                self.worksheet.cell(row=row, column=10, value=invoice_amount)
                self.period_totals[4] += invoice_amount
                print("91-120 DAYS")
            else:
                self.worksheet.cell(row=row, column=10, value=0)

            if days_overdue > 120:
                self.worksheet.cell(row=row, column=11, value=invoice_amount)
                self.period_totals[5] += invoice_amount
                print("120+ DAYS")
            else:
                self.worksheet.cell(row=row, column=11, value=0)
                
            row += 1
        
        self.worksheet.cell(row=row, column=2, value="Invoice Total")
        self.worksheet.cell(row=row, column=3, value=self.invoice_total)
        
        for i in range(len(self.period_totals)):
            self.worksheet.cell(row=row, column=6+i, value=self.period_totals[i])

        column_letter = get_column_letter(4)
        for cell in self.worksheet[column_letter]:
            cell.number_format = "dd/mm/yyyy"

        self.worksheet.column_dimensions[column_letter].auto_size = True

        self.file_name = "ARAR.xlsx"
        self.workbook.save(self.file_name)
    
    def load_invoices(self):
        # load invoices into dictionary
        print("loading invoices")
        self.workbook = openpyxl.load_workbook(filename="ARAR.xlsx")
        self.worksheet = self.workbook.active
        
        # re-do entire section below
        i = 0
        for row in self.worksheet.iter_rows(values_only=True):
            self.customer_name = row[0]
            self.invoice_number = row[1]
            self.amount_due = row[2]
            self.due_date = row[3]

            # make try catch statements for typecasting each variable
            if not self.customer_name:
                continue

            try:
                self.customer_name = str(self.customer_name)
            except:
                continue

            try:
                self.invoice_number = int(self.invoice_number)
            except:
                print("Invoice number type: ")
                print(type(self.invoice_number))
                print(self.invoice_number.split())
                continue

            try:
                self.amount_due = float(self.amount_due)
            except:
                continue

            new_invoice = {
                "customer_name": self.customer_name,
                "amount": self.amount_due,
                "due_date": self.due_date
            }

            self.invoices[self.invoice_number] = new_invoice

            print()
            print("Customer Name:")
            print(self.customer_name)
            print("Invoice Number:")
            print(self.invoice_number)
            print("Amount Due:")
            print(self.amount_due)
            print("Due Date:")
            print(self.due_date)
        print(self.invoices)


        

    def create_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "ARAR"

        if not os.path.exists("ARAR.xlsx"):
            # List of column headers
            self.headers = ["Customer Name", "Invoice Number", "Invoice Amount", "Due Date", "Days Overdue", "<0 days", "0-30 days", "31-60 days", "61-90 days", "91-120 days", "Over 120 days"]

            # Loop through the headers and set the values in the cells
            for index, header in enumerate(self.headers):
                self.worksheet.cell(row=1, column=index+1, value=header)
            
            column_letter = get_column_letter(4)
            for cell in self.worksheet[column_letter]:
                cell.number_format = "dd/mm/yyyy"
            
            self.worksheet.column_dimensions[column_letter].auto_size = True

            self.file_name = "ARAR.xlsx"
            self.workbook.save(self.file_name)

    def add_invoice(self):
        self.invoice_number = self.invoice_number_entry_1.get()
        self.customer_name = self.name_entry_1.get()
        self.due_date = self.date_entry_1.get()
        self.amount_due = self.amount_entry_1.get()

        if not self.invoice_number or not self.customer_name or not self.due_date or not self.amount_due:
            mb.showerror("Add invoice: Empty Fields", "Please make sure there are no empty fields before attempting to add a new invoice")
            return
        
        try:
            self.invoice_number = int(self.invoice_number)
        except:
            mb.showerror("Add invoice: Invalid Invoice Number", "Please enter a valid invoice number. Omit all nondigit characters in the input field")
            return
        
        if self.invoice_number <= 0:
            mb.showerror("Add invoice: Invoice number is equal to 0", "Please enter a valid invoice number greater than 0. The value entered must be greater than 0")
            return
        
        if not self.customer_name:
            mb.showerror("Add invoice: Customer Name is Empty", "Please enter a valid customer name. The input field must not be empty.")
            return
        
        try:
            self.due_date = datetime.strptime(self.due_date, "%m/%d/%Y")
            self.days_overdue = (self.due_date - datetime.today()).days
        except ValueError:
            mb.showerror("Add invoice: Invalid Due Date", "Please enter a valid date before today following the format: mm/dd/yyyy")
            return

        try:
            self.amount_due = float(self.amount_due)
        except ValueError:
            mb.showerror("Add invoice: Invalid Amount Due", "Please enter a valid amount. Omit all nondigit characters in the input field")
            return

        if self.amount_due <= 0:
            mb.showerror("Add invoice: Amount due is equal to 0", "Please enter an amount greater than 0. The value entered must be greater than 0")
            return

        for invoice_number, invoice in self.invoices.items():
            if self.invoice_number == invoice_number:
                mb.showerror("Add invoice: Invoice number already taken", "Please enter a valid invoice number that hasn't already been taken up.")
                return

        new_invoice = {
            "customer_name": self.customer_name,
            "amount": self.amount_due,
            "due_date": self.due_date
        }

        self.invoices[self.invoice_number] = new_invoice
        self.update_invoices()

        mb.showinfo("Add invoice: Successfully Added Invoice", "Invoice successfully added.")

        return

    def delete_invoice(self):
        # perform input validation:
            # nothing is blank
            # check if invoice number is int and positive
        # check if invoice exists in ARAR
            # if invoice number is the only input: use that criteria only
        # open a popup window to confirm if user wants to delete invoice 
            # display all valid data about that invoice
            # if user confirms, delete that invoice 
            # regardless, clear entries afterwards

            
        self.invoice_number = self.invoice_number_entry_2.get()
        self.customer_name = self.name_entry_2.get()
        self.due_date = self.date_entry_2.get()
        self.amount_due = self.amount_entry_2.get()

        try:
            self.invoice_number = int(self.invoice_number)
        except:
            mb.showerror("Delete invoice: Invalid Invoice Number", "Please enter a valid invoice number. Omit all nondigit characters in the input field")
            return

        if self.invoice_number:
            try:
                self.invoice_number = int(self.invoice_number)
            except:
                mb.showerror("Delete invoice: Invalid Invoice Number", "Please enter a valid invoice number. Omit all nondigit characters in the input field")
                return
            
            if self.invoice_number <= 0:
                mb.showerror("Delete invoice: Invoice number is equal to 0", "Please enter a valid invoice number greater than 0. The value entered must be greater than 0")
                return

        if self.customer_name:
            try:
                self.due_date = datetime.strptime(self.due_date, "%m/%d/%Y")
                self.days_overdue = (self.due_date - datetime.today()).days
            except ValueError:
                mb.showerror("Delete invoice: Invalid Due Date", "Please enter a valid date before today following the format: mm/dd/yyyy")
                return

        if self.amount_due:
            try:
                self.amount_due = float(self.amount_due)
            except ValueError:
                mb.showerror("Delete invoice: Invalid Amount Due", "Please enter a valid amount. Omit all nondigit characters in the input field")
                return

            if self.amount_due <= 0:
                mb.showerror("Delete invoice: Amount due is equal to 0", "Please enter an amount greater than 0. The value entered must be greater than 0")
                return
        
        found = False

        for invoice_number, invoice in self.invoices.items():
            if self.invoice_number == invoice_number:
                if self.customer_name and self.customer_name is not invoice['customer_name']:
                    mb.showerror("Delete invoice: Customer Name Not Found", "Please enter a valid customer name. An invoice with that customer name cannot be found within the ARAR report")
                    return
                if self.due_date and self.due_date is not invoice['due_date']:
                    mb.showerror("Delete invoice: Due Date Not Found", "Please enter a valid due date. An invoice with that due date cannot be found within the ARAR report")
                    return
                if self.amount_due and self.amount_due is not invoice['amount']:
                    mb.showerror("Delete invoice: Amount Not Found", "Please enter a valid amount due. An invoice with that amount due cannot be found within the ARAR report")
                    return

                found = True

        if not found:
            mb.showerror("Delete invoice: Invoice Number Not Found", "Please enter a valid invoice number. Please check if the invoice number you entered can be found on the spreadsheet")
            return

        del self.invoices[self.invoice_number]
        self.update_invoices()
        mb.showinfo("Delete invoice: Successfully Removed Invoice", "Invoice successfully removed.")
        return
    
    def view_invoices(self):
        return

    def make_user_interface(self):
        self.user_interface_frame = Frame(self, width=self.max_width, height=self.max_height)
        self.user_interface_frame.pack()

        self.title_label = Label(self.user_interface_frame, text="ARAR Maker", font=("Courier", 30, "bold"))
        self.title_label.grid(row=0, column=1, pady=5, columnspan=2)
        
        # add invoice button

        self.add_invoice_label = Label(self.user_interface_frame, text="Add an invoice", font=("Arial", 20, "bold"))
        self.add_invoice_label.grid(row=1, column=0, pady=5, columnspan=2)

        self.invoice_number_label_1 = Label(self.user_interface_frame, text="Invoice #")
        self.invoice_number_label_1.grid(row=2, column=0, pady=5)
        self.invoice_number_entry_1 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.invoice_number_entry_1.grid(row=3, column=0, pady=5, padx=7)

        self.name_label_1 = Label(self.user_interface_frame, text="Customer Name")
        self.name_label_1.grid(row=2, column=1, pady=5)
        self.name_entry_1 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.name_entry_1.grid(row=3, column=1, pady=5, padx=7)

        self.date_label_1 = Label(self.user_interface_frame, text="Due Date (mm/dd/yyyy)",)
        self.date_label_1.grid(row=2, column=2, pady=5)
        self.date_entry_1 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.date_entry_1.grid(row=3, column=2, pady=5, padx=7)

        self.amount_label_1 = Label(self.user_interface_frame, text="Amount",)
        self.amount_label_1.grid(row=2, column=3, pady=5)
        self.amount_entry_1 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.amount_entry_1.grid(row=3, column=3, pady=5, padx=7)

        # put command here
        self.add_invoice_button = Button(self.user_interface_frame, text="Add Invoice", font=("Courier", 15, "bold"), command=self.add_invoice)
        self.add_invoice_button.grid(row=1, column=2, columnspan=2, pady=5)

        # delete invoice feature

        self.delete_invoice_label = Label(self.user_interface_frame, text="Delete an Invoice", font=("Arial", 20, "bold"))
        self.delete_invoice_label.grid(row=5, column=0, pady=5, columnspan=2)
        
        self.invoice_number_label_2 = Label(self.user_interface_frame, text="Invoice #")
        self.invoice_number_label_2.grid(row=6, column=0, pady=5)
        self.invoice_number_entry_2 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.invoice_number_entry_2.grid(row=7, column=0, pady=5, padx=7)

        self.name_label_2 = Label(self.user_interface_frame, text="Customer Name (Optional)")
        self.name_label_2.grid(row=6, column=1, pady=5)
        self.name_entry_2 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.name_entry_2.grid(row=7, column=1, pady=5, padx=7)

        self.date_label_2 = Label(self.user_interface_frame, text="Due Date (mm/dd/yyyy) (Optional)",)
        self.date_label_2.grid(row=6, column=2, pady=5)
        self.date_entry_2 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.date_entry_2.grid(row=7, column=2, pady=5, padx=7)

        self.amount_label_2 = Label(self.user_interface_frame, text="Amount (Optional)",)
        self.amount_label_2.grid(row=6, column=3, pady=5)
        self.amount_entry_2 = Entry(self.user_interface_frame, width=30, borderwidth=5)
        self.amount_entry_2.grid(row=7, column=3, pady=5, padx=7)

        self.delete_invoice_button = Button(self.user_interface_frame, text="Delete Invoice", font=("Courier", 15, "bold"), command=self.delete_invoice)
        self.delete_invoice_button.grid(row=5, column=2, columnspan=2, pady=5)

        # listbox and view invoices button
        
        self.item_listbox_label = Label(self.user_interface_frame, text="AR Aging Report Overview", font=("Arial", 13))
        self.item_listbox_label.grid(row=9, column=1, columnspan=2, pady=5)

        self.item_listbox = Listbox(self.user_interface_frame, width=65, height=12, selectmode="SINGLE", borderwidth=5)
        self.item_listbox.grid(row=10, column=1, columnspan=2, pady=5)

        self.view_invoices_button = Button(self.user_interface_frame, text="View AR Aging Report", font=("Courier", 15, "bold"), command=lambda: os.system("start excel.exe ARAR.xlsx"))
        self.view_invoices_button.grid(row=11, column=1, columnspan=2, pady=5)



if __name__ == "__main__":
    arar_maker = ARARMaker()