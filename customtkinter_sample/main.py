import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from customtkinter import *
import customtkinter as ctk
from tkinter import messagebox as mb

class ARARMaker(ctk.CTk):
    def __init__(self):
        super().__init__()

        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("dark-blue")  
        self.max_width = 900
        self.max_height = 575

        self.title("ARAR Maker V0.2")
        self.geometry(str(self.max_width) + "x" + str(self.max_height))
        self.resizable(False, False)

        self.invoices = {}
        self.period_totals = [0, 0, 0, 0, 0, 0]
        self.invoice_total = 0

        self.make_user_interface()
        self.create_workbook()
        # self.load_invoices()
        self.update_invoices()
        self.update_statistics()
        self.mainloop()

    def update_statistics(self):
        return

    def update_invoices(self):
        # test me
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "ARAR"

        self.headers = ["Customer Name", "Invoice Number", "Invoice Amount", "Due Date", "Days Overdue", "<0 days", "0-30 days", "31-60 days", "61-90 days", "91-120 days", "Over 120 days"]

        # Loop through the headers and set the values in the cells
        for index, header in enumerate(self.headers):
            self.worksheet.cell(row=1, column=index+1, value=header)
        
        row = 2
        for invoice_number, invoice in self.invoices.items():
            invoice_amount = float(invoice["amount"])
            customer_name = invoice["customer_name"]
            due_date = invoice["due_date"]
            print(due_date)
            print(datetime.today())
            days_overdue = (due_date - datetime.today()).days
            print("DAYS OVERDUE: " + str(days_overdue))

            self.invoice_total += invoice_amount

            self.worksheet.cell(row=row, column=1, value=customer_name)
            self.worksheet.cell(row=row, column=2, value=invoice_number)
            self.worksheet.cell(row=row, column=3, value=invoice_amount)
            self.worksheet.cell(row=row, column=4, value=due_date)
            self.worksheet.cell(row=row, column=5, value=days_overdue)

            if days_overdue < 0:
                self.worksheet.cell(row=row, column=6, value=invoice_amount)
                self.period_totals[0] += invoice_amount
                print("<0 DAYS")
            else:
                self.worksheet.cell(row=row, column=6, value=0)

            if days_overdue >= 0 and days_overdue <= 30:
                self.worksheet.cell(row=row, column=7, value=invoice_amount)
                self.period_totals[1] += invoice_amount
                print("0-30 DAYS")
            else:
                self.worksheet.cell(row=row, column=7, value=0)

            if days_overdue >= 31 and days_overdue <= 60:
                self.worksheet.cell(row=row, column=8, value=invoice_amount)
                self.period_totals[2] += invoice_amount
                print("31-60 DAYS")
            else:
                self.worksheet.cell(row=row, column=8, value=0)

            if days_overdue >= 61 and days_overdue <= 90:
                self.worksheet.cell(row=row, column=9, value=invoice_amount)
                self.period_totals[3] += invoice_amount
                print("61-90 DAYS")
            else:
                self.worksheet.cell(row=row, column=9, value=0)

            if days_overdue >= 91 and days_overdue <= 120:
                self.worksheet.cell(row=row, column=10, value=invoice_amount)
                self.period_totals[4] += invoice_amount
                print("91-120 DAYS")
            else:
                self.worksheet.cell(row=row, column=10, value=0)

            if days_overdue > 120:
                self.worksheet.cell(row=row, column=11, value=invoice_amount)
                self.period_totals[5] += invoice_amount
                print("120+ DAYS")
            else:
                self.worksheet.cell(row=row, column=11, value=0)
                
            row += 1
        
        self.worksheet.cell(row=row, column=2, value="Invoice Total")
        self.worksheet.cell(row=row, column=3, value=self.invoice_total)
        
        for i in range(len(self.period_totals)):
            self.worksheet.cell(row=row, column=6+i, value=self.period_totals[i])

        column_letter = get_column_letter(4)
        for cell in self.worksheet[column_letter]:
            cell.number_format = "dd/mm/yyyy"

        self.worksheet.column_dimensions[column_letter].auto_size = True

        self.file_name = "ARAR.xlsx"
        self.workbook.save(self.file_name)
        return

    def load_invoices(self): 
        print("loading invoices")
        self.workbook = openpyxl.load_workbook(filename="ARAR.xlsx")
        self.worksheet = self.workbook.active
        
        # re-do entire section below
        i = 0
        for row in self.worksheet.iter_rows(values_only=True):
            self.customer_name = row[0]
            self.invoice_number = row[1]
            self.amount_due = row[2]
            self.due_date = row[3]

            # make try catch statements for typecasting each variable
            if not self.customer_name:
                continue

            try:
                self.customer_name = str(self.customer_name)
            except:
                continue

            try:
                self.invoice_number = int(self.invoice_number)
            except:
                print("Invoice number type: ")
                print(type(self.invoice_number))
                print(self.invoice_number.split())
                continue

            try:
                self.amount_due = float(self.amount_due)
            except:
                continue

            new_invoice = {
                "customer_name": self.customer_name,
                "amount": self.amount_due,
                "due_date": self.due_date
            }

            self.invoices[self.invoice_number] = new_invoice


    def create_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "ARAR"

        if not os.path.exists("ARAR.xlsx"):
            # List of column headers
            self.headers = ["Customer Name", "Invoice Number", "Invoice Amount", "Due Date", "Days Overdue", "<0 days", "0-30 days", "31-60 days", "61-90 days", "91-120 days", "Over 120 days"]

            # Loop through the headers and set the values in the cells
            for index, header in enumerate(self.headers):
                self.worksheet.cell(row=1, column=index+1, value=header)
            
            column_letter = get_column_letter(4)
            for cell in self.worksheet[column_letter]:
                cell.number_format = "dd/mm/yyyy"
            
            self.worksheet.column_dimensions[column_letter].auto_size = True

            self.file_name = "ARAR.xlsx"
            self.workbook.save(self.file_name)

    def add_invoice(self):
        return

    def delete_invoice(self):
        return

    def make_user_interface(self):
        self.user_interface_frame = CTkFrame(self, width=self.max_width, height=self.max_height)
        self.user_interface_frame.pack()

        self.title_label = CTkLabel(self.user_interface_frame, text="ARAR Maker", font=("Courier", 30))
        self.title_label.grid(row=0, column=1, pady=5, columnspan=2)
        
        # add invoice CTkButton

        self.add_invoice_label = CTkLabel(self.user_interface_frame, text="Add an invoice", font=("Lucida Console", 20))
        self.add_invoice_label.grid(row=1, column=0, pady=5, columnspan=2)

        self.invoice_number_label_1 = CTkLabel(self.user_interface_frame, text="Invoice #")
        self.invoice_number_label_1.grid(row=2, column=0, pady=5)
        self.invoice_number_entry_1 = CTkEntry(self.user_interface_frame, width=210)
        self.invoice_number_entry_1.grid(row=3, column=0, pady=5, padx=7)

        self.name_label_1 = CTkLabel(self.user_interface_frame, text="Customer Name")
        self.name_label_1.grid(row=2, column=1, pady=5)
        self.name_entry_1 = CTkEntry(self.user_interface_frame, width=210)
        self.name_entry_1.grid(row=3, column=1, pady=5, padx=7)

        self.date_label_1 = CTkLabel(self.user_interface_frame, text="Due Date (mm/dd/yyyy)",)
        self.date_label_1.grid(row=2, column=2, pady=5)
        self.date_entry_1 = CTkEntry(self.user_interface_frame, width=210)
        self.date_entry_1.grid(row=3, column=2, pady=5, padx=7)

        self.amount_label_1 = CTkLabel(self.user_interface_frame, text="Amount",)
        self.amount_label_1.grid(row=2, column=3, pady=5)
        self.amount_entry_1 = CTkEntry(self.user_interface_frame, width=210)
        self.amount_entry_1.grid(row=3, column=3, pady=5, padx=7)

        # put command here #565b5e
        self.add_invoice_button = CTkButton(self.user_interface_frame, text="Add Invoice", font=("Lucida Console", 15), command=self.add_invoice, fg_color="#343638", border_color="#565b5e")
        self.add_invoice_button.grid(row=1, column=2, columnspan=2, pady=5)

        # delete invoice feature

        self.delete_invoice_label = CTkLabel(self.user_interface_frame, text="Delete an Invoice", font=("Lucida Console", 20))
        self.delete_invoice_label.grid(row=5, column=0, pady=5, columnspan=2)
        
        self.invoice_number_label_2 = CTkLabel(self.user_interface_frame, text="Invoice #")
        self.invoice_number_label_2.grid(row=6, column=0, pady=5)
        self.invoice_number_entry_2 = CTkEntry(self.user_interface_frame, width=210)
        self.invoice_number_entry_2.grid(row=7, column=0, pady=5, padx=7)

        self.name_label_2 = CTkLabel(self.user_interface_frame, text="Customer Name (Optional)")
        self.name_label_2.grid(row=6, column=1, pady=5)
        self.name_entry_2 = CTkEntry(self.user_interface_frame, width=210)
        self.name_entry_2.grid(row=7, column=1, pady=5, padx=7)

        self.date_label_2 = CTkLabel(self.user_interface_frame, text="Due Date (mm/dd/yyyy) (Optional)",)
        self.date_label_2.grid(row=6, column=2, pady=5)
        self.date_entry_2 = CTkEntry(self.user_interface_frame, width=210)
        self.date_entry_2.grid(row=7, column=2, pady=5, padx=7)

        self.amount_label_2 = CTkLabel(self.user_interface_frame, text="Amount (Optional)",)
        self.amount_label_2.grid(row=6, column=3, pady=5)
        self.amount_entry_2 = CTkEntry(self.user_interface_frame, width=210)
        self.amount_entry_2.grid(row=7, column=3, pady=5, padx=7)

        self.delete_invoice_button = CTkButton(self.user_interface_frame, text="Delete Invoice", font=("Courier", 15), command=self.delete_invoice, fg_color="#343638", border_color="#565b5e")
        self.delete_invoice_button.grid(row=5, column=2, columnspan=2, pady=5)

        # listbox and view invoices CTkButton
        
        self.item_listbox_label = CTkLabel(self.user_interface_frame, text="AR Aging Report Overview", font=("Lucida Console", 13))
        self.item_listbox_label.grid(row=9, column=1, columnspan=2, pady=5)

        self.item_listbox = CTkTextbox(self.user_interface_frame, width=300, height=215)
        self.item_listbox.grid(row=10, column=1, columnspan=2, pady=5)

        self.view_invoices_button = CTkButton(self.user_interface_frame, text="View AR Aging Report", font=("Courier", 15), command=lambda: os.system("start excel.exe ARAR.xlsx"), fg_color="#343638", border_color="#565b5e")
        self.view_invoices_button.grid(row=11, column=1, columnspan=2, pady=5)

if __name__ == "__main__":
    arar_maker = ARARMaker()