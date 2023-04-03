import openpyxl
from datetime import datetime
from os import system

# create workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = "AR Aging Report " + str(datetime.today())

# header/title rwo
worksheet['A1'] = "Customer Name"
worksheet['B1'] = "Invoice Number"
worksheet['C1'] = "Invoice Amount"
worksheet['D1'] = "Due Date"
worksheet['E1'] = "Days Overdue"
worksheet['F1'] = "0-30 days"
worksheet['G1'] = "31-60 days"
worksheet['H1'] = "61-90 days"
worksheet['I1'] = "91-120 days"
worksheet['J1'] = "Over 120 days"

invoices = {}
invoice_total = 0 
period_totals = [0, 0, 0, 0, 0]

print("Welcome user to console-based ARAR-maker V1!")

filename = input("To begin, please enter the filename you wish to save the report as (entering no name will make 'arar.xlsx' the default filename): ")

if filename:
    filename = f"{filename}.xlsx"
else:
    filename = "arar.xlsx"

while True:
    customer_name = input("Please enter the name of the customer: ")
    
    while True:
        try:
            amount_due = input("Please enter amount due by customer: ")
            float(amount_due)
            break
        except ValueError:
            print('Please enter a valid amount.')



    while True:
        try:
            due_date = datetime.strptime(input("Please enter the payement's due date (format: mm/dd/yyyy): "), "%m/%d/%Y")
            days_overdue = (due_date - datetime.today()).days
            if days_overdue < 0:
                print("Please enter a valid date before today following the format: mm/dd/yyyy")
            else:
                break
        except ValueError:
            print("Please enter a valid date in the following format: mm/dd/yyyy")
        
    print("Customer Info: ")
    print(f"Customer's name: {customer_name}")
    print(f"Amount Due: {str(amount_due)}")
    print(f"Date Due: {due_date}")
    print(f"Days overdue: {days_overdue}")
    
    confirm = input("\nConfirm addition to AR aging report? (y/n): ").lower()

    while True:
        if confirm == 'y':
            print("Adding customer to AR aging report...")
            invoice_number = len(invoices) + 1
            new_entry = {
                'Amount': int(amount_due),
                'DueDate': due_date,
            }

            if customer_name not in invoices:
                invoices[customer_name] = {}
            invoices[customer_name][invoice_number] = new_entry
            break
        elif confirm == 'n':
            break
        else:
            confirm = input("Please enter a valid input (y/n): ")
    
    confirm = input("\nAdd another customer to the AR aging report? (y/n): ").lower()

    while True:
        if confirm == 'y':
            break
        elif confirm == 'n':
            break
        else:
            confirm = input("Please enter a valid input (y/n): ")

    if confirm == 'n':
        break

row = 2
for customer, invoice in invoices.items():
    for invoice_number, invoice_details in invoice.items():
        invoice_amount = invoice_details['Amount']
        due_date = invoice_details['DueDate']
        days_overdue = (due_date - datetime.today()).days

        worksheet.cell(row=row, column=1, value=customer)
        worksheet.cell(row=row, column=2, value=invoice_number)
        worksheet.cell(row=row, column=3, value=invoice_amount)
        worksheet.cell(row=row, column=4, value=due_date.strftime("%m/%d/%y"))
        worksheet.cell(row=row, column=5, value=days_overdue)

        invoice_total += int(invoice_amount)

        if days_overdue / 30 <= 1:
            worksheet.cell(row=row, column=6, value=invoice_amount)
            period_totals[0] += invoice_amount
        else:
            worksheet.cell(row=row, column=6, value=0)

        if days_overdue / 30 <= 2 and days_overdue / 30 > 1:
            worksheet.cell(row=row, column=7, value=invoice_amount)
            period_totals[1] += invoice_amount
        else:
            worksheet.cell(row=row, column=7, value=0)

        if days_overdue / 30 <= 3 and days_overdue / 30 >  2:
            worksheet.cell(row=row, column=8, value=invoice_amount)
            period_totals[2] += invoice_amount
        else:
            worksheet.cell(row=row, column=8, value=0)

        if days_overdue / 30 <= 4 and days_overdue / 30 >  3:
            worksheet.cell(row=row, column=9, value=invoice_amount)
            period_totals[3] += invoice_amount
        else:
            worksheet.cell(row=row, column=9, value=0)

        if days_overdue / 30 > 4:
            worksheet.cell(row=row, column=10, value=invoice_amount)
            period_totals[4] += invoice_amount
        else:
            worksheet.cell(row=row, column=10, value=0)
            
        row += 1

worksheet.cell(row=row, column=2, value="Invoice Total")
worksheet.cell(row=row, column=3, value=invoice_total)

for i in range(len(period_totals)):
    worksheet.cell(row=row, column=6+i, value=period_totals[i])


workbook.save(f"{filename}")

while True:
    view = input("Do you wish to run Excel and view the created file? (Y/N): ").lower()
    if view == 'y':
        system(f"start excel.exe {filename}")
        break
    elif view == 'n':
        print("See you soon!")
        break
    else:
        print("Please enter a valid response.")


